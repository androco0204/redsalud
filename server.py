"""
server.py - Servidor local RedSalud
Uso: python server.py
Luego abre: http://localhost:8080
"""
import http.server, json, os, sys, io, base64
from urllib.parse import urlparse, parse_qs
from datetime import date
import warnings
warnings.filterwarnings('ignore')

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

TEMPLATE = None
TEMPLATE_PATH = None

def find_template():
    candidates = [
        'Formato_Supervisión_Contractual_V2_310126___1_.xlsx',
        'Formato_Supervision_Contractual_V2.xlsx',
        'Formato_Supervision.xlsx',
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    import glob
    found = glob.glob('Formato_Super*.xlsx')
    return found[0] if found else None

def write_cell(ws, row, col, value, wrap=False):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            row, col = mr.min_row, mr.min_col
            break
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if wrap and cell.alignment:
        cell.alignment = Alignment(
            wrap_text=True,
            horizontal=cell.alignment.horizontal,
            vertical=cell.alignment.vertical or 'top'
        )

def generar_excel(datos):
    global TEMPLATE, TEMPLATE_PATH
    if not TEMPLATE_PATH:
        TEMPLATE_PATH = find_template()
    if not TEMPLATE_PATH:
        raise FileNotFoundError("No se encontró el template Excel")

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb['Formato']
    p = dados = datos.get('prestador', {})
    hab = datos.get('habilitacion', {})
    glosa = datos.get('glosa', {})
    pqr = datos.get('pqr', {})
    hoy = date.today().strftime('%d/%m/%Y')
    periodo = datos.get('periodo', '')

    # Limpiar fórmulas VLOOKUP
    for addr in ['E8','D14','G14','K14','B16','E16','H16','J16']:
        try: ws[addr].value = ''
        except: pass

    write_cell(ws,4,4,hoy); write_cell(ws,4,8,periodo)
    write_cell(ws,8,2,p.get('nit','')); write_cell(ws,8,4,p.get('nit',''))
    write_cell(ws,8,5,p.get('nombre','')); write_cell(ws,8,8,p.get('nombre',''))
    write_cell(ws,8,11,p.get('nombre','')); write_cell(ws,8,13,p.get('regional',''))
    write_cell(ws,10,4,p.get('representante_legal',''))
    write_cell(ws,10,6,'NIT'); write_cell(ws,10,7,p.get('nit',''))
    write_cell(ws,10,10,p.get('telefonos',p.get('telefono',''))); write_cell(ws,10,13,p.get('email',p.get('correo','')))
    write_cell(ws,14,2,p.get('contrato','')); write_cell(ws,14,3,p.get('contrato',''))
    write_cell(ws,14,4,p.get('descripcion_plan','REGIMEN CONTRIBUTIVO'))
    write_cell(ws,14,8,'Mensual'); write_cell(ws,14,11,p.get('fecha_inicio',''))
    write_cell(ws,16,2,p.get('modelo','')); write_cell(ws,16,5,'24')
    write_cell(ws,16,8,'Si'); write_cell(ws,16,10,'Si')
    write_cell(ws,20,2,p.get('especialidad','Prestación de servicios de salud'),wrap=True)

    # Habilitación
    hab_si=hab.get('habilitados',0); hab_tot=hab.get('total',0)
    if hab_tot:
        pct=round(hab_si/hab_tot*100)
        write_cell(ws,24,10,f'Numerador: {hab_si}\nDenominador: {hab_tot}\nResultado: {pct}%',wrap=True)
        write_cell(ws,24,12,'X' if pct==100 else ''); write_cell(ws,24,13,'X' if pct<100 else ''); write_cell(ws,24,14,'')
        write_cell(ws,24,15,hab.get('observacion',''))
    else:
        write_cell(ws,24,14,'X')

    for row in range(25,39):
        if ws.cell(row=row,column=2).value:
            write_cell(ws,row,12,''); write_cell(ws,row,13,''); write_cell(ws,row,14,'X')

    if pqr.get('total'):
        write_cell(ws,28,10,f'Total PQR período: {pqr["total"]}',wrap=True)
        write_cell(ws,28,12,'X' if pqr['total']<20 else ''); write_cell(ws,28,13,'X' if pqr['total']>=20 else ''); write_cell(ws,28,14,'')

    sub=glosa.get('subtotal',0); glos=glosa.get('glosado',0)
    if sub:
        pg=glos/sub*100
        write_cell(ws,37,10,f'% Cumplimiento Indicadores:\n{pg:.1f}%',wrap=True)
        write_cell(ws,37,12,'X' if pg<15 else ''); write_cell(ws,37,13,'X' if pg>=15 else ''); write_cell(ws,37,14,'')
        write_cell(ws,58,10,f'Resultado: {pg:.1f}%')
        write_cell(ws,58,12,'X' if pg<15 else ''); write_cell(ws,58,13,'X' if pg>=15 else ''); write_cell(ws,58,14,'')
    for row in [45,46,47,48]: write_cell(ws,row,14,'X')
    if not sub:
        write_cell(ws,56,14,'X'); write_cell(ws,57,14,'X'); write_cell(ws,58,14,'X')

    cumple = hab_tot>0 and hab_si==hab_tot and (not sub or glos/sub<0.15)
    write_cell(ws,64,14,'Sí' if cumple else 'Requiere evaluación integral')
    write_cell(ws,73,2,f'Versión 3  |  Generado: {hoy}')

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

class Handler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, format, *args): pass  # silenciar logs

    def do_POST(self):
        if self.path == '/generar-supervision':
            length = int(self.headers.get('Content-Length',0))
            body = self.rfile.read(length)
            try:
                datos = json.loads(body)
                excel_bytes = generar_excel(datos)
                nombre = datos.get('prestador',{}).get('nombre','prestador')
                nombre_file = ''.join(c if c.isalnum() or c in ' _-' else '' for c in nombre).strip()[:30].replace(' ','_')
                filename = f"Supervision_{nombre_file}_{date.today().strftime('%Y-%m-%d')}.xlsx"
                self.send_response(200)
                self.send_header('Content-Type','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                self.send_header('Content-Disposition',f'attachment; filename="{filename}"')
                self.send_header('Content-Length',str(len(excel_bytes)))
                self.send_header('Access-Control-Allow-Origin','*')
                self.end_headers()
                self.wfile.write(excel_bytes)
            except Exception as e:
                self.send_response(500)
                self.send_header('Content-Type','application/json')
                self.send_header('Access-Control-Allow-Origin','*')
                self.end_headers()
                self.wfile.write(json.dumps({'error':str(e)}).encode())
        else:
            self.send_response(404); self.end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin','*')
        self.send_header('Access-Control-Allow-Methods','POST,GET,OPTIONS')
        self.send_header('Access-Control-Allow-Headers','Content-Type')
        self.end_headers()

    def do_GET(self):
        if self.path == '/' or self.path == '/index.html':
            self.path = '/index.html'
        return super().do_GET()

PORT = 8080
print(f"\n✓ RedSalud corriendo en http://localhost:{PORT}")
print(f"  Template Excel: {find_template() or '⚠ No encontrado — pon el template en esta carpeta'}")
print("  Ctrl+C para detener\n")

os.chdir(os.path.dirname(os.path.abspath(__file__)))
httpd = http.server.HTTPServer(('', PORT), Handler)
httpd.serve_forever()
